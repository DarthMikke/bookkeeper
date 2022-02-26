import re
from datetime import datetime, timedelta
from openpyxl import load_workbook


class Transaction:
    """
    Properties:
    - date: ISO date string or ISO datetime string
    - payee: str – payee's name
    - amount: float
    """
    patterns = {
        "Gnu På Cementen": "Gnu På Cementen",
        "Www.bloomsbury.com": "Bloomsbury",
        "Lh Brennin": "Brenningen kafé",
        "Coop Prix": "Coop Prix",
        "Klassekampen": "Klassekampen",
        "Håkon Daglivare Jelsagt. 5": "Joker",
        "Flytoget": "Flytoget",
        "Clas Ohlson AS": "Clas Ohlson",
        "Vitus Apotek": "Vitus",
        "Vitusapoteket": "Vitus",
        "Fosenlinjen AS": "AtB",
        "Vinmonopolet": "Vinmonopolet",
        "Ths Engros AS Rennebu": "Ths Engros",
        "Amfora Design": "Amfora Design",
        "Spotify": "Spotify",
        "Comfort Hotel": "Comfort Hotel",
        "Burger Kin": "Burger King",
        "Far East Takeaway Stavanger": "Far East Takeaway",
        "7 - Eleven": "7 - Eleven",
        "Sellanraa": "Sellanraa",
        "Good Omens": "Good Omens",
        "Nettbuss": "Nettbuss",
        "Teezily": "Teezily",
        "Atb AS": "AtB",
        "Checkpoint Club Nedre Strand": "Checkpoint Charlie",
        "Beverly": "Beverly",
        "Narvesen": "Narvesen",
        "Norwegian.no": "Norwegian",
        "Østerlie Kunst": "Østerlie Kunst",
        "Point": "Point",
        "Lemon AS": "Lemon",
        "Itunes.com/bill": "Apple",
        "Overføring Innland  [0-9]+ (.*)": "{0}",
        "Facebk": "Facebook",
        "Mobil Billettering App": "Entur",
        "Lønn  [0-9]+ (.*)": "{0}",
        "Foodcourt Sola": "Foodcourt",
        "Ebok.no": "Ebok.no",
        "Flybussen": "Flybussen",
        "Extra": "Extra",
        "Eplehuset": "Eplehuset",
        "Scandic": "Scandic",
        "Reservert transaksjon": "Reservert beløp",
    }
    transaction_patterns = [
        'Varekjøp\s(.*)\sDato',
        'Visa\s[0-9]+\s(.*)',
    ]

    def __init__(self, date: datetime, original_payee_string: str, amount: float, match=True):
        self.date = date
        self.amount = amount
        self.original_payee_string = original_payee_string
        self.payee = self.match(original_payee_string) if match else original_payee_string

    def match(self, original_payee_string: str):
        matched_payee = None
        for pattern in Transaction.patterns.keys():
            matches = re.search(re.compile(pattern), original_payee_string)
            if matches is None:
                continue

            matched_payee = Transaction.patterns[pattern].format(*matches.groups())
            if type(matched_payee) is str:
                matched_payee = matched_payee.strip()
                return matched_payee

        for pattern in Transaction.transaction_patterns:
            matches = re.search(re.compile(pattern), original_payee_string)
            if matches is None:
                continue
            if len(matches.groups()) > 0:
                matched_payee = matches.groups()[0].strip()
                break

        if matched_payee is None:
            matched_payee = original_payee_string
        return matched_payee

    def as_list(self):
        if self.payee is None:
            return None
        return [self.date.date().isoformat(), self.payee, self.amount]


def parse_transactions(filepath: str, first: datetime or None, last: datetime or None):
    """
    Parse transactions from a bank statement.
    @param filepath: path to a .xlsx file
    @param first: first day to parse
    @param last: last day to parse
    @return: list of Transaction or None objects. None means that the row
    was invalid or outside the desired time range.
    """
    # TODO: Check if file exists
    wb = load_workbook(filepath)
    # TODO: Check if the workbook has any sheets
    # TODO: Throw error if the first sheet not in the correct format
    sheet = wb.sheetnames[0]
    ws = wb[sheet]

    transactions = [_parse_single_transaction(x, first, last) for x in ws.rows]
    transactions = [x for x in transactions if x is not None]
    return transactions


def _match_datetime(description, year):
    datetime_pattern = re.compile("([0-9]{2})\.([0-9]{2}) kl\. ([0-9]{2})\.([0-9]{2})")
    matches = re.search(datetime_pattern, description)
    if matches is None:
        return None
    matches = tuple((int(x) for x in matches.groups()))
    return datetime(year, matches[1], matches[0], matches[2], matches[3])


def _parse_single_transaction(row, first: datetime, last: datetime):
    last = datetime(last.year, last.month, last.day) + timedelta(1)
    if row[1].value == "Forklaring":
        return None

    dt = row[0].value
    if dt is None:
        return None

    matched_dt = _match_datetime(row[1].value, dt.year)
    if matched_dt is not None:
        dt = matched_dt

    # Return None if outside of desired time range
    if type(dt) is not datetime:
        return None
    if dt < first or dt > last:
        return None

    try:
        transaction = Transaction(
            dt,
            row[1].value,
            -float(row[3].value) if row[3].value is not None else float(row[4].value)
        )
    except Exception as e:
        print(e)
        return None
    return transaction
