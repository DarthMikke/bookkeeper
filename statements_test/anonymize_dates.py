# -*- coding: utf-8 -*-

import re
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook

wb = load_workbook("./transaksjonliste_orig.xlsx")
sheet = wb.sheetnames[0]
ws = wb[sheet]

start = datetime(2016, 12, 1)
pattern = re.compile("[0-9]{2}\.[0-9]{2} kl\. [0-9]{2}\.[0-9]{2}")

for row in ws.rows:
    if row[1].value == "Forklaring":
        continue
    print('|'.join([str(x.value) for x in row]), end='')
    print('\t=>\t', end='')
    transaction = row[1].value
    has_date = re.search(pattern, transaction) is not None
    new_date = (start + timedelta(random.randint(0, 41), hours=random.randint(4, 23), minutes=random.randint(0, 59)))
    new_transaction_string = re.sub(pattern, new_date.strftime("%d.%m kl. %H.%M"), transaction)
    row[0].value = new_date
    row[1].value = new_transaction_string
    if row[2].value is not None:
        row[2].value = new_date \
            if (new_date.month, new_date.day) not in [(12, 31), (1, 1), (1, 2)]\
            else datetime(2017, 1, 3)
    print('|'.join([str(x.value) for x in row[0:1]]))

w.save("transaksjonsliste.xlsx")